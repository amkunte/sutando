#!/usr/bin/env python3
"""Parse the finance sheet (.xlsx export) into state/latest-snapshot.json.

Deterministic — reads the xlsx with openpyxl, NOT the read_file_content text
rep (which truncates the 1,800-row `daily` tab to ~191 rows and made an earlier
analysis wrongly think the tab was dead). The `daily` tab is the live
net-worth time-series + the authoritative day-over-day change; the `Dashboard`
tab supplies composition/concentration/FIRE/milestone.

Usage:
    extract_sheet.py /path/to/finance.xlsx
Writes skills/finance-radar/state/latest-snapshot.json and prints a summary.
Requires openpyxl (pip install --user openpyxl).
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

from common import latest_path

try:
    import openpyxl
except ImportError:
    print("extract_sheet: openpyxl not installed — run: pip install --user openpyxl",
          file=sys.stderr)
    raise SystemExit(3)


def _num(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def _sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower() == name.lower():
            return wb[s]
    return None


def parse_daily(wb) -> dict:
    """Latest net worth (Total), day-over-day change ($ + %), and last-7 trend."""
    ws = _sheet(wb, "daily")
    out = {}
    if ws is None:
        return out
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return out
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

    def col(name):
        return header.index(name) if name in header else None
    ci_date, ci_total = col("date"), col("total")
    ci_chg, ci_chgp = col("change"), col("change %")
    if ci_total is None:
        return out
    # data rows that have a numeric Total, in order
    data = [r for r in rows[1:] if r and ci_total < len(r) and _num(r[ci_total]) is not None]
    if not data:
        return out
    last = data[-1]
    out["net_worth"] = _num(last[ci_total])
    if ci_date is not None and ci_date < len(last) and last[ci_date] is not None:
        d = last[ci_date]
        out["as_of"] = d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d)
    if ci_chg is not None and ci_chg < len(last):
        out["day_change"] = _num(last[ci_chg])
    if ci_chgp is not None and ci_chgp < len(last):
        p = _num(last[ci_chgp])
        out["day_change_pct"] = (p * 100.0) if p is not None else None  # stored as fraction
    out["trend_7d"] = [_num(r[ci_total]) for r in data[-7:]]
    return out


def _find_label(ws_rows, text):
    """Return (r, c) of the first cell whose stripped text equals `text`
    (case-insensitive), else None."""
    t = text.strip().lower()
    for r, row in enumerate(ws_rows):
        for c, v in enumerate(row):
            if isinstance(v, str) and v.strip().lower() == t:
                return (r, c)
    return None


def _label_contains(ws_rows, sub):
    s = sub.strip().lower()
    for r, row in enumerate(ws_rows):
        for c, v in enumerate(row):
            if isinstance(v, str) and s in v.strip().lower():
                return (r, c)
    return None


def parse_dashboard(wb) -> dict:
    ws = _sheet(wb, "Dashboard")
    out = {}
    if ws is None:
        return out
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    def at(r, c):
        return rows[r][c] if 0 <= r < len(rows) and 0 <= c < len(rows[r]) else None

    # ATH ("Max  Net Worth" — note double space) → nearest numeric below the label
    loc = _label_contains(rows, "max") and _label_contains(rows, "net worth")
    loc = next((( r, c) for (r, row) in enumerate(rows) for (c, v) in enumerate(row)
                if isinstance(v, str) and "max" in v.lower() and "net worth" in v.lower()), None)
    if loc:
        for dr in range(1, 4):
            n = _num(at(loc[0] + dr, loc[1]))
            if n is not None:
                out["ath"] = n
                break

    # Mag-7 concentration: label 'Concentration' → first fraction in same row to the right
    loc = _find_label(rows, "Concentration")
    if loc:
        for c in range(loc[1] + 1, loc[1] + 4):
            n = _num(at(loc[0], c))
            if n is not None:
                out.setdefault("mag7", {})["concentration_pct"] = n * 100.0 if n <= 1 else n
                break

    # Liquid Cash: label → value directly below
    loc = _find_label(rows, "Liquid Cash")
    if loc:
        for dr in range(1, 3):
            n = _num(at(loc[0] + dr, loc[1]))
            if n is not None:
                out["liquid_cash"] = n
                break

    # FIRE net worth: label contains 'FIRE Net Worth' → value one column right
    loc = _label_contains(rows, "FIRE Net Worth")
    if loc:
        for dc in range(1, 3):
            n = _num(at(loc[0], loc[1] + dc))
            if n is not None:
                out["fire_net_worth"] = n
                break

    # Asset-class split: the 'Live' row in the Net-Worth-Trend block.
    # Header row has Cash/Liabilities/Real Estate/Side Inv/Stock/Grand Total.
    hdr = _find_label(rows, "Grand Total")
    live = _find_label(rows, "Live")
    if hdr and live:
        hr, _ = hdr
        names = {}
        for c, v in enumerate(rows[hr]):
            if isinstance(v, str) and v.strip() in (
                    "Cash", "Liabilities", "Real Estate", "Side Inv", "Stock"):
                names[c] = v.strip()
        ac = {}
        for c, label in names.items():
            n = _num(at(live[0], c))
            if n is not None:
                ac[label] = n
        if ac:
            out["asset_classes"] = ac

    # Next milestone: the milestone ladder — labels like '9M','9.5M','10M' with a
    # First-Date column; the next milestone is the first label whose date is blank.
    # Find the '# of days'/'First Date' block: labels in one col, dates in col+1.
    # Heuristic: scan for cells matching '<num>M' and a neighbor date.
    ladder = []
    for r, row in enumerate(rows):
        for c, v in enumerate(row):
            if isinstance(v, str) and v.strip().rstrip("M").replace(".", "", 1).isdigit() \
                    and v.strip().endswith("M"):
                date_cell = at(r, c + 1)
                amt = float(v.strip()[:-1]) * 1_000_000
                ladder.append((amt, date_cell))
    if ladder:
        undated = [amt for amt, d in sorted(ladder) if d in (None, "")]
        if undated:
            out["next_milestone"] = undated[0]

    return out


def main() -> int:
    if len(sys.argv) < 2:
        print("usage: extract_sheet.py <finance.xlsx>", file=sys.stderr)
        return 2
    wb = openpyxl.load_workbook(sys.argv[1], data_only=True, read_only=True)
    snap = {}
    snap.update(parse_dashboard(wb))   # composition first
    snap.update(parse_daily(wb))       # daily wins for net_worth/as_of (live source)
    # merge mag7 dict if both touched it
    if "net_worth" not in snap:
        print("extract_sheet: could not find a net worth in the daily tab", file=sys.stderr)
        return 2
    latest_path().write_text(json.dumps(snap, indent=2, ensure_ascii=False))
    print(f"wrote latest-snapshot.json: net worth {snap.get('net_worth')} "
          f"as of {snap.get('as_of')}, day Δ {snap.get('day_change')} "
          f"({snap.get('day_change_pct')}%), trend pts {len(snap.get('trend_7d') or [])}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
